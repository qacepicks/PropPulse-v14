#!/usr/bin/env python3
# prop_ev.py — PropPulse+ v2025.3
# L20-weighted projection + FantasyPros DvP + Auto position + Manual odds entry

# ===============================
# IMPORTS
# ===============================
import requests
import pandas as pd
import numpy as np
from scipy.stats import norm
import os, json, time, math
from datetime import datetime, timezone
from dvp_updater import load_dvp_data
import math
import sys
print("DEBUG: stdin.isatty() =", sys.stdin.isatty())

# ===============================
# ENHANCED DISPLAY SYSTEM
# ===============================
from display_results import (
    display_top_props,
    show_by_probability,
    show_high_confidence,
    show_by_stat_type,
    display_summary_stats,
    export_to_csv,
    export_to_markdown,
    interactive_display
)

# ===============================
# AUTO FETCH PLAYER LOGS
# ===============================
import requests
import os
import pandas as pd

def load_player_logs(player):
    """Loads local logs; if missing, auto-fetches from BallDontLie API."""
    safe_name = player.lower().replace('.', '').replace("'", "").replace(' ', '_')
    path = f"data/{safe_name}.csv"

    if os.path.exists(path):
        return pd.read_csv(path)

    # --- Auto-fetch fallback ---
    print(f"[Data] 🆕 Fetching logs for missing player: {player}")
    try:
        resp = requests.get(
            f"https://www.balldontlie.io/api/v1/players?search={player}"
        ).json()
        if not resp["data"]:
            raise ValueError("Player not found on BallDontLie")

        player_id = resp["data"][0]["id"]
        games = requests.get(
            f"https://www.balldontlie.io/api/v1/stats?player_ids[]={player_id}&per_page=100"
        ).json()["data"]

        df = pd.DataFrame([
            {
                "PTS": g.get("pts", 0),
                "REB": g.get("reb", 0),
                "AST": g.get("ast", 0),
                "FG3M": g.get("fg3m", 0),
                # defensive check for nested keys
                "game_date": g.get("game", {}).get("date", None)
            }
            for g in games
        ])

        # ensure the column always exists
        if "game_date" not in df.columns:
            df["game_date"] = None


        os.makedirs("data", exist_ok=True)
        df.to_csv(path, index=False)
        print(f"[Data] ✅ Cached new log → {path}")
        return df

    except Exception as e:
        print(f"[Data] ❌ Failed to fetch {player}: {e}")
        return pd.DataFrame(columns=["PTS","REB","AST","FG3M"])

# ===============================
# GLOBAL DEFAULTS
# ===============================
pace_mult = 1.0      # baseline neutral pace (league average)
is_home = None        # None means ignore home/away bump


# --- Import model calibration constants ---
from calibration import (
    TEMP_Z, BALANCE_BIAS, CLIP_MIN, CLIP_MAX,
    MULT_CENTER, MULT_MAX_DEV, EMP_PRIOR_K, W_EMP_MAX,
    INCLUDE_LAST_SEASON, SHRINK_TO_LEAGUE
)
# ===============================
# LOAD DVP DATA
# ===============================
dvp_data = load_dvp_data()

# ==========================================
# 🔧 Balanced Scoring & Probability System
# ==========================================
# CONFIG / CALIBRATION KNOBS
TEMP_Z = 1.45          # >1.0 flattens probabilities (less extreme 10–90%)
BALANCE_BIAS = 0.20    # pull toward 50%
CLIP_MIN, CLIP_MAX = 0.08, 0.92

# Multiplier normalization
MULT_CENTER = 1.00
MULT_MAX_DEV = 0.15

# Empirical smoothing
EMP_PRIOR_K = 20
W_EMP_MAX = 0.30

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def _cap(x, lo, hi):
    return max(lo, min(hi, x))

def normalize_multiplier(raw: float,
                         center: float = MULT_CENTER,
                         max_dev: float = MULT_MAX_DEV) -> float:
    """Normalize DvP / pace / homeaway multipliers to stay near 1.0"""
    if raw <= 0 or not math.isfinite(raw):
        return 1.0
    log_m = math.log(raw / center)
    log_m *= 0.5
    m = math.exp(log_m)
    return _cap(m, 1.0 - max_dev, 1.0 + max_dev)

def adjusted_mean(mu_base: float,
                  multipliers: dict,
                  league_mu: float | None = None,
                  shrink_to_league: float = 0.10) -> float:
    """Apply normalized multipliers & optional shrink toward league average."""
    prod = 1.0
    for k, v in (multipliers or {}).items():
        prod *= normalize_multiplier(float(v))
    mu = mu_base * prod
    if league_mu is not None and math.isfinite(league_mu):
        mu = (1.0 - shrink_to_league) * mu + shrink_to_league * league_mu
    return mu

def prob_over_from_normal(mu: float, sigma: float, line: float,
                          temp_z: float = TEMP_Z) -> float:
    """Flattened z-score probability (smaller slope -> less extreme results)."""
    eps = 1e-9
    sigma_eff = max(sigma, eps)
    z = (line - mu) / sigma_eff
    z /= max(1.0, temp_z)
    return float(1.0 - norm.cdf(z))

def smooth_empirical_prob(vals: np.ndarray, line: float, k: int = EMP_PRIOR_K) -> float:
    """Empirical hit rate vs the line with Beta prior smoothing toward 50%."""
    n = int(vals.size)
    hits = int((vals > line).sum())
    alpha = hits + 0.5 * k
    beta = (n - hits) + 0.5 * k
    return alpha / (alpha + beta)

def finalize_prob(p_raw: float,
                  balance_bias: float = BALANCE_BIAS,
                  clip_min: float = CLIP_MIN,
                  clip_max: float = CLIP_MAX) -> float:
    """Pull probabilities gently toward 50% and clip to safe range."""
    p = (1.0 - balance_bias) * p_raw + balance_bias * 0.5
    return _cap(p, clip_min, clip_max)

# ==========================================
# CORE WRAPPER — Balanced Probability
# ==========================================
def calibrated_prob_over(
    mu_base: float,
    sigma_base: float,
    line: float,
    multipliers: dict,
    recent_vals: np.ndarray,
    league_mu: float | None = None
) -> float:
    """
    Returns a balanced P(over) after:
    1) Normalizing multipliers
    2) Flattening z-score (temperature)
    3) Blending in empirical hit-rate (sample-size aware)
    4) Applying a small balance bias toward 50%
    5) Clipping to sane bounds
    """
    mu = adjusted_mean(mu_base, multipliers, league_mu=league_mu, shrink_to_league=0.10)
    p_model = prob_over_from_normal(mu, sigma_base, line, temp_z=TEMP_Z)

    p_emp = smooth_empirical_prob(np.array(recent_vals, dtype=float), line)
    n = int(len(recent_vals))
    w_emp = min(W_EMP_MAX, n / (n + EMP_PRIOR_K))
    p_blend = (1 - w_emp) * p_model + w_emp * p_emp

    return finalize_prob(p_blend, balance_bias=BALANCE_BIAS,
                         clip_min=CLIP_MIN, clip_max=CLIP_MAX)

# ==========================================
# EV / ODDS
# ==========================================
def american_to_prob(odds: int) -> float:
    return abs(odds)/(abs(odds)+100) if odds < 0 else 100/(odds+100)

def net_payout(odds: int) -> float:
    return 100/abs(odds) if odds < 0 else odds/100

def ev_per_dollar(p: float, odds: int) -> float:
    """Expected value per $1 wager"""
    return p * net_payout(odds) - (1 - p)

def ev_sportsbook(p, odds):
    return p * net_payout(odds) - (1 - p)

# ===============================
# CONFIG
# ===============================
def load_settings():
    default = {
        "default_sportsbook": "Fliff",
        "default_region": "us",
        "data_path": "data/",
        "injury_api_key": "YOUR_SPORTSDATAIO_KEY",
        "balldontlie_api_key": "YOUR_BALLDONTLIE_KEY",
        "cache_hours": 24
    }
    path = "settings.json"

    if not os.path.exists(path):
        with open(path, "w") as f:
            json.dump(default, f, indent=4)
        print("[Config] Created new settings.json.")
        return default

    with open(path, "r") as f:
        settings = json.load(f)

    for k, v in default.items():
        if k not in settings:
            settings[k] = v

    os.makedirs(settings["data_path"], exist_ok=True)
    return settings

def get_bdl(endpoint, params=None, settings=None, timeout=10):
    """
    BallDontLie V1 API helper with correct authentication for paid tier.
    """
    base_url = "https://api.balldontlie.io/v1"
    url = base_url + endpoint
    
    headers = {
        "User-Agent": "PropPulse/1.0",
        "Accept": "application/json"
    }
    
    # ✅ FIXED: Paid API uses Bearer token format
    if settings:
        api_key = settings.get("balldontlie_api_key")
        if api_key and api_key != "YOUR_BALLDONTLIE_KEY":
            # Try Bearer format first (most common for paid APIs)
            headers["Authorization"] = f"Bearer {api_key}"
    
    params = params or {}
    
    try:
        r = requests.get(url, headers=headers, params=params, timeout=timeout)
        
        print(f"[BDL] GET {endpoint} with params={params}")
        print(f"[BDL] Status: {r.status_code}")
        
        # Detailed error handling
        if r.status_code == 401:
            print(f"[BDL] ❌ 401 Unauthorized")
            api_key = settings.get("balldontlie_api_key") if settings else None
            if api_key and api_key != "YOUR_BALLDONTLIE_KEY":
                print(f"[BDL] Using key: {api_key[:10]}...{api_key[-4:]}")
                print(f"[BDL] Key length: {len(api_key)} characters")
            else:
                print("[BDL] ❌ No valid API key configured in settings.json")
            print("[BDL] 💡 Tip: Check if your API key is correct at https://app.balldontlie.io/")
            return None
            
        elif r.status_code == 429:
            print("[BDL] ⚠️ Rate limited - waiting 2s...")
            # Check rate limit headers
            if 'X-RateLimit-Remaining' in r.headers:
                print(f"[BDL] Requests remaining: {r.headers['X-RateLimit-Remaining']}")
            time.sleep(2)
            return None
            
        elif r.status_code == 404:
            print(f"[BDL] ⚠️ 404 Not Found: {url}")
            return None
            
        elif r.status_code == 403:
            print(f"[BDL] ❌ 403 Forbidden - Your API key may not have access to this endpoint")
            return None
            
        elif r.status_code == 200:
            data = r.json()
            result_count = len(data.get('data', []))
            print(f"[BDL] ✅ Success - returned {result_count} records")
            
            # Debug: Show what we got
            if result_count > 0 and 'data' in data:
                first_item = data['data'][0]
                print(f"[BDL] First result preview: {first_item.get('first_name', '')} {first_item.get('last_name', '')}")
            
            return data
        else:
            print(f"[BDL] ⚠️ Unexpected status code: {r.status_code}")
            print(f"[BDL] Response: {r.text[:200]}")
            return None
        
    except requests.exceptions.Timeout:
        print(f"[BDL] ⚠️ Request timeout after {timeout}s")
        return None
    except requests.exceptions.RequestException as e:
        print(f"[BDL] ⚠️ Request failed: {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"[BDL] ❌ Failed to parse JSON response: {e}")
        print(f"[BDL] Raw response: {r.text[:200]}")
        return None

# ===============================
# ✅ FIXED: POSITION DETECTION WITH BETTER INFERENCE
# ===============================
def get_player_position_auto(player_name, df_logs=None, settings=None):
    """
    Automatically fetches player position using BallDontLie V1 API.
    Falls back to improved stat-based inference if API fails.
    """
    # Try BallDontLie V1 API first
    try:
        print(f"[Position] 🔍 Searching BallDontLie for '{player_name}'...")
        
        # BallDontLie search works better with last name only
        last_name = player_name.split()[-1]
        print(f"[Position] Searching by last name: '{last_name}'")
        
        data = get_bdl("/players", {"search": last_name}, settings)
        
        if data and "data" in data and len(data["data"]) > 0:
            print(f"[Position] Found {len(data['data'])} matching player(s)")
            
            # Try exact match first
            for player in data["data"]:
                full_name = f"{player.get('first_name', '')} {player.get('last_name', '')}".strip()
                print(f"[Position] Checking: {full_name}")
                if full_name.lower() == player_name.lower():
                    pos = player.get("position", "").strip().upper()
                    if pos:
                        pos = normalize_position(pos)
                        print(f"[Position] ✅ BallDontLie V1 (exact match) → {pos}")
                        return pos
            
            # Fall back to first result
            pos = data["data"][0].get("position", "").strip().upper()
            if pos:
                pos = normalize_position(pos)
                first_match = f"{data['data'][0].get('first_name', '')} {data['data'][0].get('last_name', '')}"
                print(f"[Position] ✅ BallDontLie V1 (using first result: {first_match}) → {pos}")
                return pos
        else:
            print(f"[Position] ⚠️ BallDontLie returned no results for '{last_name}'")
        
    except Exception as e:
        print(f"[Position] ⚠️ BallDontLie error: {e}")
    
    # Enhanced fallback: Infer from stats with better thresholds
    print(f"[Position] 🔍 Using enhanced stat-based inference...")
    if df_logs is not None and len(df_logs) > 0:
        return infer_position_from_stats(df_logs, player_name)
    
    print("[Position] ⚠️ No data available, defaulting to SF")
    return "SF"

def normalize_position(pos):
    """Normalize position abbreviations to standard 5 positions"""
    pos = pos.upper().strip()
    
    # Map variations to standard positions
    position_map = {
        "G": "SG",
        "G-F": "SF",
        "F": "SF", 
        "F-G": "SF",
        "F-C": "PF",
        "C-F": "C"
    }
    
    return position_map.get(pos, pos) if pos in position_map else pos

def infer_position_from_stats(df_logs, player_name=""):
    """
    Improved position inference using multiple statistical indicators.
    """
    def avg(col):
        if col not in df_logs.columns:
            return 0.0
        return pd.to_numeric(df_logs[col], errors="coerce").fillna(0).mean()
    
    a_pts = avg("PTS")
    a_reb = avg("REB")
    a_ast = avg("AST")
    a_fg3 = avg("FG3M")
    
    # Calculate ratios for better classification
    ast_reb_ratio = a_ast / a_reb if a_reb > 0 else 0
    
    print(f"[Position] Stats: PTS={a_pts:.1f} REB={a_reb:.1f} AST={a_ast:.1f} 3PM={a_fg3:.1f}")
    
    # Enhanced position logic
    # Point Guard: High assists, lower rebounds
    if a_ast >= 6.5:
        print(f"[Position] 🔍 Inferred PG (high AST: {a_ast:.1f})")
        return "PG"
    
    # Center: High rebounds, low assists, low 3-pointers
    if a_reb >= 9 and a_fg3 < 1.0:
        print(f"[Position] 🔍 Inferred C (high REB: {a_reb:.1f}, low 3PM)")
        return "C"
    
    # Power Forward: Good rebounds, moderate scoring
    if a_reb >= 7.5 and a_ast < 5.5:
        print(f"[Position] 🔍 Inferred PF (REB: {a_reb:.1f}, AST: {a_ast:.1f})")
        return "PF"
    
    # Small Forward: Balanced stats (THIS WAS THE MISSING CASE FOR SCOTTIE BARNES)
    # Typical SF: 15-20 pts, 5-8 reb, 3-5 ast
    if 4.5 <= a_reb <= 8 and 3 <= a_ast <= 6 and a_pts >= 12:
        print(f"[Position] 🔍 Inferred SF (balanced: PTS={a_pts:.1f}, REB={a_reb:.1f}, AST={a_ast:.1f})")
        return "SF"
    
    # Shooting Guard: Moderate assists, lower rebounds, good scoring
    if a_ast >= 3 and a_reb < 5.5 and (a_fg3 >= 1.5 or a_pts >= 15):
        print(f"[Position] 🔍 Inferred SG (AST: {a_ast:.1f}, REB: {a_reb:.1f})")
        return "SG"
    
    # Default based on rebounds if nothing else matches
    if a_reb >= 7:
        print(f"[Position] 🔍 Default to PF (high REB: {a_reb:.1f})")
        return "PF"
    elif a_reb >= 5:
        print(f"[Position] 🔍 Default to SF (moderate REB: {a_reb:.1f})")
        return "SF"
    else:
        print(f"[Position] 🔍 Default to SG")
        return "SG"

# ===============================
# ✅ FIXED: OPPONENT DETECTION WITH BETTER ERROR HANDLING
# ===============================
def get_upcoming_opponent_abbr(player_name, settings=None):
    """
    Uses BallDontLie V1 to pull the player's next opponent.
    Improved search logic and error handling.
    """
    try:
        # Get player info with better search
        print(f"[Schedule] 🔍 Looking up next game for {player_name}...")
        
        # BallDontLie search works better with last name only
        last_name = player_name.split()[-1]
        print(f"[Schedule] Searching by last name: '{last_name}'")
        
        player_data = get_bdl("/players", {"search": last_name}, settings)
        
        if not player_data or "data" not in player_data or not player_data["data"]:
            print(f"[Schedule] ⚠️ Player not found via API")
            return None
        
        # Try to find exact match
        player = None
        for p in player_data["data"]:
            full_name = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip()
            if full_name.lower() == player_name.lower():
                player = p
                print(f"[Schedule] ✅ Found exact match: {full_name}")
                break
        
        # Fall back to first result if no exact match
        if not player:
            player = player_data["data"][0]
            fallback_name = f"{player.get('first_name', '')} {player.get('last_name', '')}"
            print(f"[Schedule] Using first result: {fallback_name}")
        
        team = player.get("team", {})
        team_id = team.get("id")
        team_abbr = team.get("abbreviation", "UNK")
        
        if not team_id:
            print("[Schedule] ⚠️ No team ID found")
            return None
        
        print(f"[Schedule] Player team: {team_abbr} (ID: {team_id})")
        
        # Get current season
        today = datetime.now(timezone.utc)
        season = today.year if today.month >= 10 else today.year - 1
        
        # Get team's games
        games_data = get_bdl("/games", {
            "seasons[]": season,
            "team_ids[]": team_id,
            "per_page": 100
        }, settings)
        
        if not games_data or "data" not in games_data:
            print("[Schedule] ⚠️ No games data returned")
            return None
        
        games = games_data["data"]
        print(f"[Schedule] Found {len(games)} games for season {season}")
        
        # Find next game
        today_date = today.date()
        future_games = []
        
        for game in games:
            try:
                game_date_str = game.get("date", "")
                game_date = datetime.fromisoformat(game_date_str.replace("Z", "+00:00")).date()
                
                if game_date >= today_date:
                    future_games.append((game_date, game))
            except Exception as e:
                continue
        
        if not future_games:
            print("[Schedule] ⚠️ No upcoming games found")
            return None
        
        # Sort by date and get next game
        future_games.sort(key=lambda x: x[0])
        next_game_date, next_game = future_games[0]
        
        # Determine opponent
        home_team = next_game.get("home_team", {})
        visitor_team = next_game.get("visitor_team", {})
        
        if home_team.get("id") == team_id:
            opp_abbr = visitor_team.get("abbreviation")
            location = "vs"
        else:
            opp_abbr = home_team.get("abbreviation")
            location = "@"
        
        if opp_abbr:
            print(f"[Schedule] ✅ Next game: {next_game_date} {location} {opp_abbr}")
            return opp_abbr
        
        return None
        
    except Exception as e:
        print(f"[Schedule] ⚠️ Error: {e}")
        import traceback
        traceback.print_exc()
        return None

# ===============================
# INJURY STATUS
# ===============================
def get_injury_status(player_name, api_key):
    if not api_key or "YOUR_SPORTSDATAIO_KEY" in api_key:
        return None
    try:
        url = "https://api.sportsdata.io/v4/nba/scores/json/Players"
        r = requests.get(url, headers={"Ocp-Apim-Subscription-Key": api_key}, timeout=8)
        if r.status_code != 200:
            return None
        for p in r.json():
            if player_name.lower() in p.get("Name", "").lower():
                return p.get("InjuryStatus", None)
    except Exception:
        return None
    return None

# ===============================
# UNIVERSAL PLAYER LOG FETCHER (AUTO FALLBACK)
# ===============================
def fetch_player_logs(player_name, save_dir="data/", settings=None, include_last_season=True):
    """
    Unified fetcher: tries BallDontLie V1 first, then Basketball Reference.
    Returns clean DataFrame or None.
    """
    import requests
    import pandas as pd
    import os
    from datetime import datetime
    from bs4 import BeautifulSoup
    import time

    # === 1. Try BallDontLie V1 API first ===
    try:
        print(f"[BDL] Trying V1 API for {player_name}...")
        last_name = player_name.split()[-1]
        
        player_data = get_bdl("/players", {"search": last_name}, settings)
        
        if player_data and "data" in player_data and len(player_data["data"]) > 0:
            # Find matching player
            player = None
            for p in player_data["data"]:
                full_name = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip()
                if full_name.lower() == player_name.lower():
                    player = p
                    break
            
            if not player:
                player = player_data["data"][0]  # Use first result
            
            player_id = player.get("id")
            print(f"[BDL] ✅ Found {player_name} (ID {player_id})")
            
            # Get current season stats
            # BallDontLie uses the START year (2024-25 season = 2024)
            today = datetime.now()
            if today.month >= 10:
                season = today.year - 1  # Nov 2025 → use 2024 (for 2024-25 season)
            elif today.month <= 6:
                season = today.year - 1  # Jan-June 2025 → use 2024
            else:
                season = today.year  # July-Sept → use current year
            print(f"[BDL] Calculated season: {season} (for {season}-{season+1} NBA season)")
            
            # V1 uses different endpoint structure
            stats_data = get_bdl("/stats", {
                "player_ids[]": player_id,
                "seasons[]": season,
                "per_page": 100
            }, settings)
            
            if stats_data and "data" in stats_data:
                games = stats_data["data"]
                print(f"[BDL] Retrieved {len(games)} games for season {season}")
                
                if games:
                    rows = []
                    for g in games:
                        mins_raw = g.get("min", "0")
                        try:
                            if ":" in str(mins_raw):
                                mins = float(mins_raw.split(":")[0])
                            else:
                                mins = float(mins_raw or 0)
                        except:
                            mins = 0.0
                        
                        rows.append({
                            "GAME_ID": g.get("game", {}).get("id", ""),
                            "DATE": g.get("game", {}).get("date", ""),
                            "PTS": g.get("pts", 0),
                            "REB": g.get("reb", 0),
                            "AST": g.get("ast", 0),
                            "FG3M": g.get("fg3m", 0),
                            "MIN": mins
                        })
                    
                    df = pd.DataFrame(rows)
                    df = df[df["MIN"] > 0]
                    
                    if len(df) > 0:
                        path = os.path.join(save_dir, f"{player_name.replace(' ', '_')}.csv")
                        os.makedirs(save_dir, exist_ok=True)
                        df.to_csv(path, index=False)
                        print(f"[Save] ✅ {len(df)} games saved → {path}")
                        return df
        
        print(f"[BDL] ⚠️ No data found via V1 API")
    except Exception as e:
        print(f"[BDL] ❌ V1 API error: {e}")

    # === 2. Fallback: Basketball Reference (improved) ===
    print("[Fallback] Trying Basketball Reference...")

    def bbref_name_format(name):
        """Generate Basketball Reference player ID slug"""
        name = name.lower().replace(".", "").replace("'", "").replace("-", "")
        parts = name.split()
        if len(parts) < 2:
            return None
        last, first = parts[-1], parts[0]
        # Format: last5 + first2 + 01
        return f"{last[:5]}{first[:2]}01"

    slug = bbref_name_format(player_name)
    if not slug:
        print(f"[BBRef] ❌ Invalid name format: {player_name}")
        return None
    
    # Get first letter of last name for URL
    last_name = player_name.split()[-1]
    first_letter = last_name[0].lower()
    
    # ✅ FIXED: Use correct year for BBRef
    # BBRef uses the ENDING year (2024-25 season = 2025)
    year = datetime.now().year
    if datetime.now().month < 10:  # Before October
        year = year  # Already the correct ending year
    else:  # October or later
        year = year + 1  # Use next year
    
    rows = []
    
    seasons_to_try = [year, year - 1] if include_last_season else [year]
    
    for yr in seasons_to_try:
        url = f"https://www.basketball-reference.com/players/{first_letter}/{slug}/gamelog/{yr}"
        print(f"[BBRef] Fetching {yr} season: {url}")
        
        try:
            # ✅ ENHANCED: Better headers and user agent rotation
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none",
                "Cache-Control": "max-age=0"
            }
            
            # Add respectful delay
            time.sleep(3)
            
            html = requests.get(url, headers=headers, timeout=15)
            
            if html.status_code == 403:
                print(f"[BBRef] ⚠️ 403 Forbidden for {yr} - site is blocking automated access")
                print(f"[BBRef] 💡 Try using a VPN or manually download the data")
                continue
            elif html.status_code == 404:
                print(f"[BBRef] ⚠️ 404 Not Found for {yr}")
                continue
            elif html.status_code != 200:
                print(f"[BBRef] ⚠️ Status {html.status_code} for {yr}")
                continue
            
            soup = BeautifulSoup(html.text, "html.parser")
            table = soup.find("table", {"id": "pgl_basic"})
            
            if not table:
                print(f"[BBRef] ⚠️ No game log table found for {yr}")
                continue
            
            print(f"[BBRef] ✅ Found game log table for {yr}")
            
            # Parse table rows
            for tr in table.find_all("tr"):
                if not tr.find("td"):
                    continue
                
                tds = [td.text.strip() for td in tr.find_all("td")]
                
                if len(tds) < 27:
                    continue
                
                try:
                    pts = float(tds[26]) if tds[26] else 0
                    reb = float(tds[22]) if tds[22] else 0
                    ast = float(tds[23]) if tds[23] else 0
                    fg3m = float(tds[11]) if tds[11] else 0
                    
                    mins_str = tds[6]
                    if mins_str and mins_str != "Did Not Play":
                        try:
                            if ":" in mins_str:
                                m, s = mins_str.split(":")
                                mins = int(m) + int(s) / 60.0
                            else:
                                mins = float(mins_str)
                        except:
                            mins = 0
                    else:
                        mins = 0
                    
                    if mins > 0:
                        rows.append({
                            "PTS": pts,
                            "REB": reb,
                            "AST": ast,
                            "FG3M": fg3m,
                            "MIN": mins
                        })
                except (ValueError, IndexError):
                    continue
            
            print(f"[BBRef] ✅ Parsed {len([r for r in rows if r])} games from {yr}")
            
        except requests.exceptions.Timeout:
            print(f"[BBRef] ⚠️ Timeout fetching {yr}")
        except Exception as e:
            print(f"[BBRef] ❌ Error fetching {yr}: {e}")

    if not rows:
        print(f"[BBRef] ❌ No data found for {player_name}")
        return None

    df = pd.DataFrame(rows)
    df = df[df["MIN"] > 0]
    path = os.path.join(save_dir, f"{player_name.replace(' ', '_')}.csv")
    os.makedirs(save_dir, exist_ok=True)
    df.to_csv(path, index=False)
    print(f"[Save] ✅ {len(df)} games saved → {path}")
    print("[Source] ✅ Using Basketball Reference data")
    return df

# ===============================
# L20-WEIGHTED MODEL
# ===============================
def l20_weighted_mean(vals: pd.Series) -> float:
    if len(vals) == 0:
        return 0.0
    season_mean = float(pd.to_numeric(vals, errors="coerce").fillna(0).mean())
    last20 = pd.to_numeric(vals.tail(20), errors="coerce").fillna(0)
    l20_mean = float(last20.mean()) if len(last20) > 0 else season_mean
    return 0.60 * l20_mean + 0.40 * season_mean


# ===============================
# PROBABILITY CALIBRATION
# ===============================
import math

def prob_calibrate(p: float, T: float = 1.15, b: float = 0.0, shrink: float = 0.20) -> float:
    """
    Temperature-calibrates model probability and shrinks toward 0.5.
    T > 1 lowers confidence; T < 1 raises it. b shifts bias in logit space.
    shrink in [0,1]: 0 = no shrink, 1 = force to 0.5.
    """
    p = max(1e-6, min(1 - 1e-6, float(p)))
    logit = math.log(p / (1 - p))
    logit = (logit / T) + b
    q = 1.0 / (1.0 + math.exp(-logit))
    return 0.5 + (1.0 - shrink) * (q - 0.5)


# ===============================
# ✅ FIXED: GRADE PROBABILITIES
# ===============================
def grade_probabilities(df, stat_col, line, proj_mins, avg_mins, injury_status=None, dvp_mult=1.0):
    """
    Fixed: Removed player_name parameter to match app.py calls
    """
    if stat_col not in df.columns:
        if stat_col == "REB+AST":
            df["REB+AST"] = df["REB"] + df["AST"]
        elif stat_col == "PRA":
            df["PRA"] = df["PTS"] + df["REB"] + df["AST"]
        else:
            raise KeyError(f"Missing stat {stat_col}")

    vals = pd.to_numeric(df[stat_col], errors="coerce").fillna(0.0)
    n = len(vals)
    std = float(vals.std(ddof=0)) if n > 1 else 1.0
    mean = l20_weighted_mean(vals)

    # === Base Minute Scaling ===
    mean *= (proj_mins / avg_mins) if avg_mins > 0 else 1.0

    # === Injury & Minute Dampener ===
    if injury_status and str(injury_status).lower() not in ["active", "probable"]:
        mean *= 0.9  # injury penalty (reduce 10%)
    elif proj_mins < avg_mins * 0.8:
        mean *= 0.9  # fewer minutes than average

    # === Opponent Pace Adjustment ===
    pace_mult = globals().get("pace_mult", 1.0)  # safe fallback
    mean *= pace_mult
    print(f"[Model] Pace multiplier: {pace_mult:.3f}")

    # === Home / Away Adjustment ===
    is_home = globals().get("is_home", None)
    if is_home is True:
        mean *= 1.03  # 3% bump at home
    elif is_home is False:
        mean *= 0.97  # 3% drop on road

    # === Rebounding Calibration Fix ===
    if stat_col in ["REB", "REB+AST", "PRA"]:
        dvp_mult *= 0.85  # reduce overbias on rebound-heavy stats

    # === Category-specific Volatility (σ scaling) ===
    sigma_scale = {"PTS":1.00, "REB":0.85, "AST":0.90, "PRA":0.95, "REB+AST":0.88, "FG3M":1.10}
    std *= sigma_scale.get(stat_col, 1.0)
    print(f"[Model] Volatility scale ({stat_col}): {sigma_scale.get(stat_col,1.0)}")

    # Apply DvP
    mean *= float(dvp_mult)
    print(f"[Model] DvP (adjusted): {dvp_mult:.3f} | Adjusted mean: {mean:.2f}")

    # --- Run Probability Model ---
    recent_vals = np.array(vals, dtype=float)
    multipliers = {
        "dvp": dvp_mult,
        "pace": pace_mult,
        "ha": 1.0,
        "h2h": 1.0
    }

    p_final = calibrated_prob_over(
        mu_base=mean,
        sigma_base=std,
        line=line,
        multipliers=multipliers,
        recent_vals=recent_vals,
        league_mu=None
    )
    try:
        # === Probability Calibration Post-Process ===
        T = 1.15  # can be tuned per stat if desired
        p_cal = prob_calibrate(p_final, T=T, b=0.0, shrink=0.20)
        print(f"[Calib] raw={p_final:.3f} → calibrated={p_cal:.3f} (T={T}, shrink=0.20)")

        return p_cal, n, mean

    except Exception as e:
        # --- Safety fallback: ensures model never crashes ---
        print(f"[Model] ❌ Failsafe triggered: {e}")
        safe_p = 0.5  # neutral probability
        safe_n = n if 'n' in locals() else 0
        safe_mean = mean if 'mean' in locals() else 0.0
        return safe_p, safe_n, safe_mean


# ===============================
# DvP MULTIPLIER - FIXED FOR COMBINED STATS
# ===============================
def get_dvp_multiplier(opponent_abbr, position, stat_key):
    """
    Get DvP multiplier for a stat against an opponent.
    Handles combined stats like REB+AST and PRA by averaging component ranks.
    """
    try:
        if not opponent_abbr or not position or not stat_key:
            return 1.0
        
        # Normalize inputs
        opponent_abbr = opponent_abbr.upper()
        position = position.upper()
        stat_key = stat_key.upper()
        
        # Check if opponent exists
        if opponent_abbr not in dvp_data:
            print(f"[DvP] ⚠️ Team {opponent_abbr} not in DvP data")
            return 1.0
        
        team_dict = dvp_data[opponent_abbr]
        
        # Check if position exists
        if position not in team_dict:
            print(f"[DvP] ⚠️ Position {position} not found for {opponent_abbr}")
            return 1.0
        
        pos_dict = team_dict[position]
        
        # Handle combined stats
        if stat_key == "REB+AST":
            # Average the ranks for REB and AST
            reb_rank = pos_dict.get("REB")
            ast_rank = pos_dict.get("AST")
            
            if reb_rank is not None and ast_rank is not None:
                avg_rank = (reb_rank + ast_rank) / 2.0
                multiplier = 1.1 - (avg_rank - 1) / 300
                print(f"[DvP] {opponent_abbr} vs {position}: REB rank={reb_rank}, AST rank={ast_rank}, avg={avg_rank:.1f} → {multiplier:.3f}")
                return multiplier
            else:
                print(f"[DvP] ⚠️ Missing REB or AST data for {opponent_abbr} {position}")
                return 1.0
        
        elif stat_key == "PRA":
            # Average the ranks for PTS, REB, and AST
            pts_rank = pos_dict.get("PTS")
            reb_rank = pos_dict.get("REB")
            ast_rank = pos_dict.get("AST")
            
            if all(r is not None for r in [pts_rank, reb_rank, ast_rank]):
                avg_rank = (pts_rank + reb_rank + ast_rank) / 3.0
                multiplier = 1.1 - (avg_rank - 1) / 300
                print(f"[DvP] {opponent_abbr} vs {position}: PTS={pts_rank}, REB={reb_rank}, AST={ast_rank}, avg={avg_rank:.1f} → {multiplier:.3f}")
                return multiplier
            else:
                print(f"[DvP] ⚠️ Missing PTS/REB/AST data for {opponent_abbr} {position}")
                return 1.0
        
        # Single stat lookup
        else:
            rank = pos_dict.get(stat_key)
            
            if rank is None:
                print(f"[DvP] ⚠️ Stat {stat_key} not found for {opponent_abbr} {position}")
                return 1.0
            
            multiplier = 1.1 - (rank - 1) / 300
            print(f"[DvP] {opponent_abbr} vs {position} on {stat_key}: rank={rank} → {multiplier:.3f}")
            return multiplier
            
    except Exception as e:
        print(f"[DvP] ❌ Error calculating multiplier: {e}")
        return 1.0


# ===============================
# 🔍 DEBUG HELPER
# ===============================
def debug_projection(df, stat="REB+AST", line=12.5, player_name=""):
    """Debug helper to understand the projection breakdown"""
    
    # Create combined stat if needed
    if stat == "REB+AST" and stat not in df.columns:
        df["REB+AST"] = df["REB"] + df["AST"]
    elif stat == "PRA" and stat not in df.columns:
        df["PRA"] = df["PTS"] + df["REB"] + df["AST"]
    
    vals = pd.to_numeric(df[stat], errors="coerce").fillna(0.0)
    
    print("\n" + "="*60)
    print(f"🔍 DEBUG: {player_name} {stat} Projection Analysis")
    print("="*60)
    
    # Season stats
    season_mean = vals.mean()
    season_std = vals.std()
    season_median = vals.median()
    print(f"\n📊 Full Season Stats ({len(vals)} games):")
    print(f"   Mean: {season_mean:.2f}")
    print(f"   Median: {season_median:.2f}")
    print(f"   Std Dev: {season_std:.2f}")
    print(f"   Min: {vals.min():.1f} | Max: {vals.max():.1f}")
    
    # Last 20 games
    last20 = vals.tail(20)
    l20_mean = last20.mean()
    l20_median = last20.median()
    print(f"\n📈 Last 20 Games:")
    print(f"   Mean: {l20_mean:.2f}")
    print(f"   Median: {l20_median:.2f}")
    print(f"   Difference from season: {l20_mean - season_mean:+.2f}")
    
    # L20 weighted calculation (matching your model)
    weighted_mean = 0.60 * l20_mean + 0.40 * season_mean
    print(f"\n⚖️ L20-Weighted Mean:")
    print(f"   (60% × {l20_mean:.2f}) + (40% × {season_mean:.2f})")
    print(f"   = {weighted_mean:.2f}")
    
    # Minutes adjustment
    if "MIN" in df.columns:
        avg_mins = df["MIN"].mean()
        recent_mins = df["MIN"].tail(10).mean()
        print(f"\n⏱️ Minutes Per Game:")
        print(f"   Season Avg: {avg_mins:.1f}")
        print(f"   Last 10 Games: {recent_mins:.1f}")
    
    # Hit rate at line
    over_count = (vals > line).sum()
    hit_rate = over_count / len(vals) * 100
    print(f"\n🎯 Historical Performance vs Line {line}:")
    print(f"   Over: {over_count}/{len(vals)} ({hit_rate:.1f}%)")
    print(f"   Under: {len(vals) - over_count}/{len(vals)} ({100-hit_rate:.1f}%)")
    
    # Last 20 games hit rate
    last20_over = (last20 > line).sum()
    last20_rate = last20_over / len(last20) * 100
    print(f"   Last 20: {last20_over}/20 over ({last20_rate:.1f}%)")
    
    # Recent trend - last 10 games
    print(f"\n📉 Last 10 Games {stat}:")
    last10_vals = vals.tail(10).values
    for i, v in enumerate(last10_vals, 1):
        status = "✅ OVER" if v > line else "❌ UNDER"
        print(f"   Game -{10-i+1}: {v:.1f} {status}")
    
    # Percentile analysis
    percentile_25 = vals.quantile(0.25)
    percentile_75 = vals.quantile(0.75)
    print(f"\n📊 Distribution:")
    print(f"   25th percentile: {percentile_25:.1f}")
    print(f"   50th percentile: {season_median:.1f}")
    print(f"   75th percentile: {percentile_75:.1f}")
    print(f"   Line {line} is at {(vals <= line).sum() / len(vals) * 100:.1f}th percentile")
    
    print("="*60 + "\n")


# ===============================
# MAIN
# ===============================
def analyze_single_prop(player, stat, line, odds, settings, debug_mode=False):
    """Analyze a single prop and return results dict"""
    path = os.path.join(settings["data_path"], f"{player.replace(' ', '_')}.csv")
    need_refresh = not os.path.exists(path) or (time.time() - os.path.getmtime(path)) / 3600 > 24

    if need_refresh:
        print(f"[Data] ⏳ Refreshing logs for {player}...")
        try:
            df = fetch_player_logs(player, save_dir=settings["data_path"], settings=settings)
            if df is not None and len(df) > 0:
                print("[Source] ✅ Using BallDon'tLie data")
            else:
                raise ValueError("No data from BallDon'tLie")
        except Exception as e:
            print(f"[BDL] ⚠️ BallDon'tLie failed: {e}")
            print("[Fallback] Trying Basketball Reference instead...")
            df = fetch_player_logs(
                player,
                save_dir=settings["data_path"],
                settings=settings,
                include_last_season=INCLUDE_LAST_SEASON
            )
        
        if df is None or len(df) == 0:
            print(f"[Logs] ❌ Could not fetch logs for {player}.")
            return None
        print("[Source] ✅ Data loaded successfully")
    else:
        df = pd.read_csv(path)
        print(f"[Data] ✅ Loaded {len(df)} games for {player}")

    # Data cleaning
    if "MIN" in df.columns:
        def parse_minutes(val):
            if isinstance(val, str) and ":" in val:
                try:
                    m, s = val.split(":")
                    return int(m) + int(s)/60
                except:
                    return 0.0
            try:
                return float(val)
            except:
                return 0.0

        df["MIN"] = df["MIN"].apply(parse_minutes)
        before = len(df)
        df = df[df["MIN"] > 0]
        after = len(df)
        print(f"[Clean] Removed {before - after} DNP games ({after} remain).")

    avg_mins = df["MIN"].mean() if "MIN" in df.columns and len(df["MIN"]) > 0 else 30
    proj_mins = avg_mins

    inj = get_injury_status(player, settings.get("injury_api_key"))
    pos = get_player_position_auto(player, df_logs=df, settings=settings)
    opp = get_upcoming_opponent_abbr(player, settings=settings)
    
    dvp_mult = get_dvp_multiplier(opp, pos, stat) if opp else 1.0

    if debug_mode:
        debug_projection(df, stat=stat, line=line, player_name=player)

    try:
        p_model, n_games, proj_stat = grade_probabilities(
            df, stat, line, proj_mins, avg_mins, inj, dvp_mult
        )
    except Exception as e:
        print(f"[Model] Error: {e}")
        return None

    # 🚨 Outlier filter (skip unrealistic gaps)
    if abs(proj_stat - line) > 6:
        print(f"[Filter] ⚠️ Skipping {player} — Unrealistic gap ({proj_stat:.1f} vs line {line})")
        return None

    p_book = american_to_prob(odds)
    ev = ev_sportsbook(p_model, odds)

    return {
        "player": player,
        "stat": stat,
        "line": line,
        "odds": odds,
        "opponent": opp,
        "position": pos,
        "n_games": n_games,
        "p_model": p_model,
        "p_book": p_book,
        "projection": proj_stat,
        "ev": ev,
        "dvp_mult": dvp_mult,
        "injury": inj
    }



def batch_analyze_props(props_list, settings):
    """Analyze multiple props and return sorted by EV"""
    results = []
    for i, prop in enumerate(props_list, 1):
        print(f"\n{'='*60}")
        print(f"Analyzing prop {i}/{len(props_list)}: {prop['player']} {prop['stat']} {prop['line']}")
        print(f"{'='*60}")
        result = analyze_single_prop(
            prop['player'],
            prop['stat'],
            prop['line'],
            prop['odds'],
            settings
        )
        if result:
            results.append(result)
    results.sort(key=lambda x: x['ev'], reverse=True)
    return results

def main(player=None, stat=None, line=None, odds=None, silent=False):
    settings = load_settings()
    
    # --- PRIORITY 1: Auto-bypass menu if running from automation with a CSV ---
    import sys
    import glob
    
    # Check if CSV was passed as argument OR auto-detect
    csv_path = None
    if len(sys.argv) > 1 and sys.argv[1].endswith(".csv"):
        csv_path = sys.argv[1]
    else:
        # Auto-detect the newest CSV
        csv_files = sorted(glob.glob("auto_props_*.csv"), key=os.path.getmtime, reverse=True)
        if csv_files:
            csv_path = csv_files[0]
    
    # If we found a CSV, process it immediately
    if csv_path and os.path.exists(csv_path):
        print("🧠 PropPulse+ Model v2025.3 — Player Prop EV Analyzer")
        print("==============================\n")
        print(f"📂 Auto-detected CSV: {csv_path}\n")
        
        import pandas as pd
        try:
            props_df = pd.read_csv(csv_path)
            props = props_df.to_dict("records")
            print(f"⏳ Analyzing {len(props)} props...\n")
            results = batch_analyze_props(props, settings)
            
            if results:
                print(f"\n✅ Analysis complete! Found {len(results)} valid props")
                df = pd.DataFrame(results)
                export_results_to_excel(df)
            else:
                print("No valid results to display.")
            return results
        except Exception as e:
            print(f"❌ Error processing CSV: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    # --- PRIORITY 2: Interactive menu (only if no CSV found) ---
    print("🧠 PropPulse+ Model v2025.3 — Player Prop EV Analyzer")
    print("==============================\n")

    # --- Streamlit / Silent Mode ---
    if player and stat and line is not None and odds is not None:
        result = analyze_single_prop(player, stat, line, odds, settings)
        if not result:
            return {"Player": player, "Output": "No data returned"}

        p_model = result.get("p_model", 0)
        p_book = result.get("p_book", 0)
        ev = result.get("ev", 0)
        proj_stat = result.get("projection", 0)
        n_games = result.get("n_games", 0)

        # Return structured JSON for Streamlit
        if silent:
            return {
                "Player": player,
                "Stat": stat,
                "Line": line,
                "Odds": odds,
                "Projection": round(proj_stat, 2),
                "ModelProb%": f"{p_model * 100:.2f}%",
                "BookProb%": f"{p_book * 100:.2f}%",
                "EV_per_$1": f"{ev * 100:.1f}¢",
                "Edge%": f"{(p_model - p_book) * 100:.2f}%",
                "Opponent": result.get('opponent', 'N/A'),
                "Position": result.get('position', 'N/A'),
                "DvP": round(result.get('dvp_mult', 1.0), 3),
                "GamesAnalyzed": n_games,
                "InjuryStatus": result.get('injury', 'N/A'),
                "Result": "🟢 OVER Value" if proj_stat > line else "🔴 UNDER Value"
            }

        # Normal terminal printout
        print("\n==============================")
        print(f"📊 {player} | {stat} Line {line}")
        print(f"Games Analyzed: {n_games}")
        print(f"Model Prob:  {p_model * 100:.1f}%")
        print(f"Book Prob:   {p_book * 100:.1f}%")
        print(f"Model Projection: {proj_stat:.1f} {stat}")
        print(f"EV:          {ev * 100:.1f}¢ per $1 | {'🔥 Positive' if ev > 0 else '⚠️ Negative'}")
        print("🟢 Over Value" if proj_stat > line else "🔴 Under Value")
        print("==============================\n")
        return

    # --- Interactive CLI Mode (manual use only) ---
    mode = input("Mode: [1] Single Prop or [2] Batch Analysis (default=1): ").strip() or "1"

    if mode == "2":
        print("\n📋 BATCH MODE: Enter props one by one (blank player name when done)")
        props_list = []
        while True:
            print(f"\n--- Prop #{len(props_list) + 1} ---")
            player = input("Player name (or press Enter to finish): ").strip()
            if not player:
                break
            stat = input("Stat (PTS/REB/AST/REB+AST/PRA/FG3M): ").strip().upper()
            line = float(input("Line: "))
            odds = int(input("Odds (e.g. -110): "))
            props_list.append({
                "player": player,
                "stat": stat,
                "line": line,
                "odds": odds
            })

        if not props_list:
            print("No props entered. Exiting.")
            return

        print(f"\n⏳ Analyzing {len(props_list)} props...")
        results = batch_analyze_props(props_list, settings)
        if results:
            print(f"\n✅ Analysis complete! Found {len(results)} valid props")
            print("="*80)
            interactive_display(results)
        else:
            print("No valid results to display.")
        return

    # --- Single CLI Run ---
    player = input("Player name: ").strip()
    stat = input("Stat (PTS / REB / AST / REB+AST / PRA / FG3M): ").strip().upper()
    line = float(input("Line: "))
    odds = int(input("Sportsbook odds (e.g. -110): "))
    debug_mode = input("Enable debug mode? (y/n, default=n): ").strip().lower() == 'y'

    result = analyze_single_prop(player, stat, line, odds, settings, debug_mode)
    if not result:
        return

    p_model = result['p_model']
    p_book = result['p_book']
    ev = result['ev']
    proj_stat = result['projection']
    n_games = result['n_games']

    print("\n==============================")
    print(f"📊 {player} | {stat} Line {line}")
    print(f"Games Analyzed: {n_games}")
    print(f"Model Prob:  {p_model * 100:.1f}%")
    print(f"Book Prob:   {p_book * 100:.1f}%")
    print(f"Model Projection: {proj_stat:.1f} {stat}")
    print(f"EV:          {ev * 100:.1f}¢ per $1 | {'🔥 Positive' if ev > 0 else '⚠️ Negative'}")
    print("🟢 Over Value" if proj_stat > line else "🔴 Under Value")
    print("==============================\n")


# ===============================
# EXPORT HANDLER (PropPulse Dashboard Edition)
# ===============================
def export_results_to_excel(df):
    """Enhanced PropPulse dashboard export with color tiers, summary, and heatmap."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import ColorScaleRule
    
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    xlsx_path = os.path.join(output_dir, f"model_output_{timestamp}.xlsx")
    print("📈 Generating PropPulse Excel dashboard...\n")

    rename_map = {
        "player": "Player", "stat": "Stat", "line": "Line", "odds": "Odds",
        "p_model": "Model Prob", "p_book": "Book Prob",
        "projection": "Projection", "ev": "EV ($/1)", "dvp_mult": "DvP Mult",
        "injury": "Injury"
    }
    df.rename(columns=rename_map, inplace=True, errors="ignore")

    for col in ["Odds", "EV ($/1)", "Model Prob", "Book Prob"]:
        if col not in df.columns:
            df[col] = 0

    VALID_ODDS_RANGE = (-140, 140)
    SORT_MODE = "ev"

    df["Model Prob"] = (df.get("Model Prob", 0) * 100).round(2)
    df["Book Prob"] = (df.get("Book Prob", 0) * 100).round(2)
    df["EV ($/1)"] = df.get("EV ($/1)", 0).round(4)

    def ev_prizepicks(p_model):
        return (p_model - 0.542) * 100

    def assign_tier(ev, ev_min, ev_max):
        if ev >= ev_max * 0.75:
            return "ELITE"
        elif ev >= ev_max * 0.40:
            return "SOLID"
        elif ev <= ev_min * 0.5:
            return "FADE"
        else:
            return "NEUTRAL"

    df_filtered = df[df["Odds"].between(*VALID_ODDS_RANGE)].copy()
    df_filtered["EV (PrizePicks)"] = df_filtered["Model Prob"].apply(lambda x: ev_prizepicks(x / 100))
    df_sorted = df_filtered.sort_values(
        "Model Prob" if SORT_MODE == "prob" else "EV ($/1)", ascending=False
    )

    ev_min, ev_max = df_sorted["EV ($/1)"].min(), df_sorted["EV ($/1)"].max()
    df_sorted["Tier"] = [
        assign_tier(ev, ev_min, ev_max) if not pd.isna(ev) else ""
        for ev in df_sorted["EV ($/1)"]
    ]

    df_sorted.to_excel(xlsx_path, index=False)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.title = "ALL_PROPS"

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    color_map = {
        "ELITE": "00C0A87E",
        "SOLID": "00FFD966",
        "FADE": "00E06666",
        "NEUTRAL": "00D9D9D9"
    }

    tier_col = [c.column for c in ws[1] if c.value == "Tier"][0]
    for row in range(2, ws.max_row + 1):
        tier = ws.cell(row=row, column=tier_col).value
        if tier in color_map:
            fill = PatternFill(start_color=color_map[tier], end_color=color_map[tier], fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    ev_col = [c.column for c in ws[1] if c.value == "EV ($/1)"][0]
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"
    )
    ws.conditional_formatting.add(
        f"{ws.cell(2, ev_col).coordinate}:{ws.cell(ws.max_row, ev_col).coordinate}", rule
    )

    summary = wb.create_sheet("SUMMARY")
    summary.append(["Metric", "Value"])
    summary.append(["Timestamp", timestamp])
    summary.append(["Total Props", len(df_sorted)])
    summary.append(["Mean EV ($/1)", f"{df_sorted['EV ($/1)'].mean():.4f}"])
    summary.append(["Mean Model Prob (%)", f"{df_sorted['Model Prob'].mean():.2f}%"])
    summary.append(["Mean PP EV (%)", f"{df_sorted['EV (PrizePicks)'].mean():.2f}%"])

    for tier in ["ELITE", "SOLID", "FADE", "NEUTRAL"]:
        summary.append([f"{tier} Count", len(df_sorted[df_sorted['Tier'] == tier])])

    summary.append(["Projected ROI ($1 per prop)", f"${df_sorted['EV ($/1)'].sum():.2f}"])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(xlsx_path)
    print(f"\n✅ Excel dashboard saved: {xlsx_path}")

    try:
        if platform.system() == "Windows":
            os.startfile(xlsx_path)
        elif platform.system() == "Darwin":
            subprocess.call(["open", xlsx_path])
        else:
            subprocess.call(["xdg-open", xlsx_path])
    except Exception as e:
        print(f"⚠️ Could not auto-open Excel: {e}")


# ===============================
# BATCH ANALYSIS FUNCTION
# ===============================
def batch_analyze_props(props_list, settings):
    """Analyze multiple props and return sorted by EV"""
    results = []
    total = len(props_list)
    
    for i, prop in enumerate(props_list, 1):
        print(f"\n{'='*60}")
        print(f"Analyzing prop {i}/{total}: {prop['player']} {prop['stat']} {prop['line']}")
        print(f"{'='*60}")
        
        try:
            result = analyze_single_prop(
                prop['player'],
                prop['stat'],
                prop['line'],
                prop['odds'],
                settings
            )
            if result:
                results.append(result)
                print(f"✅ Success - EV: {result['ev']*100:.1f}¢")
            else:
                print(f"⚠️ Skipped - No valid data")
        except Exception as e:
            print(f"❌ Error analyzing {prop['player']}: {e}")
            continue
    
    results.sort(key=lambda x: x['ev'], reverse=True)
    return results


def find_latest_csv():
    """Find the most recent auto_props_*.csv file"""
    try:
        csv_files = glob.glob("auto_props_*.csv")
        if not csv_files:
            return None
        
        csv_files.sort(key=os.path.getmtime, reverse=True)
        latest = csv_files[0]
        
        age_hours = (time.time() - os.path.getmtime(latest)) / 3600
        if age_hours > 24:
            print(f"⚠️ Latest CSV is {age_hours:.1f} hours old - may be stale")
        
        return latest
    except Exception as e:
        print(f"⚠️ Error finding CSV: {e}")
        return None


def run_csv_batch_mode(csv_path):
    """Process a CSV file with props in batch mode"""
    print("🧠 PropPulse+ Model v2025.3 — Batch Mode")
    print("="*60)
    print(f"📂 Processing: {csv_path}\n")
    
    settings = load_settings()
    
    try:
        df = pd.read_csv(csv_path)
        print(f"✅ Loaded {len(df)} props from CSV")
        
        required = ['player', 'stat', 'line', 'odds']
        missing = [col for col in required if col not in df.columns]
        
        if missing:
            print(f"❌ CSV missing required columns: {missing}")
            print(f"   Found columns: {list(df.columns)}")
            return None
        
        props = df.to_dict('records')
        
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")
        import traceback
        traceback.print_exc()
        return None
    
    print(f"\n⏳ Starting analysis of {len(props)} props...\n")
    results = batch_analyze_props(props, settings)
    
    if results:
        print(f"\n✅ Analysis complete! {len(results)}/{len(props)} props analyzed successfully")
        
        results_df = pd.DataFrame(results)
        export_results_to_excel(results_df)
        
        print("\n📊 Quick Summary:")
        print(f"   Positive EV props: {len([r for r in results if r['ev'] > 0])}")
        print(f"   Average EV: {sum(r['ev'] for r in results) / len(results) * 100:.1f}¢")
        print(f"   Top EV: {max(results, key=lambda x: x['ev'])['player']} "
              f"({max(r['ev'] for r in results) * 100:.1f}¢)")
        
        return results
    else:
        print("\n❌ No valid results to export")
        return None


def main(player=None, stat=None, line=None, odds=None, silent=False):
    """Main entry point with proper CSV auto-detection"""
    
    settings = load_settings()
    
    # Check for CSV argument
    csv_path = None
    
    if len(sys.argv) > 1 and sys.argv[1].endswith('.csv'):
        csv_path = sys.argv[1]
        if not os.path.exists(csv_path):
            print(f"❌ CSV file not found: {csv_path}")
            return None
    
    elif sys.stdin.isatty():
        csv_path = find_latest_csv()
        if csv_path:
            response = input(f"📂 Found recent CSV: {csv_path}\n   Use it for batch analysis? (y/n): ").strip().lower()
            if response != 'y':
                csv_path = None
    
    if csv_path:
        return run_csv_batch_mode(csv_path)
    
    # Programmatic/silent mode
    if player and stat and line is not None and odds is not None:
        result = analyze_single_prop(player, stat, line, odds, settings)
        
        if not result:
            return {"Player": player, "Output": "No data returned"}
        
        if silent:
            return {
                "Player": player,
                "Stat": stat,
                "Line": line,
                "Odds": odds,
                "Projection": round(result['projection'], 2),
                "ModelProb%": f"{result['p_model'] * 100:.2f}%",
                "BookProb%": f"{result['p_book'] * 100:.2f}%",
                "EV_per_$1": f"{result['ev'] * 100:.1f}¢",
                "Edge%": f"{(result['p_model'] - result['p_book']) * 100:.2f}%",
                "Opponent": result.get('opponent', 'N/A'),
                "Position": result.get('position', 'N/A'),
                "DvP": round(result.get('dvp_mult', 1.0), 3),
                "GamesAnalyzed": result['n_games'],
                "InjuryStatus": result.get('injury', 'N/A'),
                "Result": "🟢 OVER Value" if result['projection'] > line else "🔴 UNDER Value"
            }
        
        print("\n" + "="*60)
        print(f"📊 {player} | {stat} Line {line}")
        print(f"Games Analyzed: {result['n_games']}")
        print(f"Model Prob:  {result['p_model'] * 100:.1f}%")
        print(f"Book Prob:   {result['p_book'] * 100:.1f}%")
        print(f"Model Projection: {result['projection']:.1f} {stat}")
        print(f"EV: {result['ev'] * 100:.1f}¢ per $1 | {'🔥 Positive' if result['ev'] > 0 else '⚠️ Negative'}")
        print("🟢 Over Value" if result['projection'] > line else "🔴 Under Value")
        print("="*60 + "\n")
        return result
    
    # Interactive CLI
    print("🧠 PropPulse+ Model v2025.3 — Player Prop EV Analyzer")
    print("="*60 + "\n")
    
    mode = input("Mode: [1] Single Prop or [2] Batch Entry (default=1): ").strip() or "1"
    
    if mode == "2":
        print("\n📋 BATCH MODE: Enter props one by one (blank player name when done)")
        props_list = []
        
        while True:
            print(f"\n--- Prop #{len(props_list) + 1} ---")
            player = input("Player name (or press Enter to finish): ").strip()
            
            if not player:
                break
            
            try:
                stat = input("Stat (PTS/REB/AST/REB+AST/PRA/FG3M): ").strip().upper()
                line = float(input("Line: "))
                odds = int(input("Odds (e.g. -110): "))
                
                props_list.append({
                    "player": player,
                    "stat": stat,
                    "line": line,
                    "odds": odds
                })
                print(f"✅ Added: {player} {stat} {line} ({odds})")
                
            except ValueError as e:
                print(f"❌ Invalid input: {e}")
                continue
        
        if not props_list:
            print("No props entered. Exiting.")
            return None
        
        print(f"\n⏳ Analyzing {len(props_list)} props...")
        results = batch_analyze_props(props_list, settings)
        
        if results:
            print(f"\n✅ Analysis complete! Found {len(results)} valid props")
            df = pd.DataFrame(results)
            export_results_to_excel(df)
            interactive_display(results)
        else:
            print("No valid results to display.")
        
        return results
    
    # Single prop mode
    player = input("Player name: ").strip()
    stat = input("Stat (PTS/REB/AST/REB+AST/PRA/FG3M): ").strip().upper()
    line = float(input("Line: "))
    odds = int(input("Sportsbook odds (e.g. -110): "))
    debug_mode = input("Enable debug mode? (y/n, default=n): ").strip().lower() == 'y'
    
    result = analyze_single_prop(player, stat, line, odds, settings, debug_mode)
    
    if not result:
        print("❌ Analysis failed")
        return None
    
    print("\n" + "="*60)
    print(f"📊 {player} | {stat} Line {line}")
    print(f"Games Analyzed: {result['n_games']}")
    print(f"Model Prob:  {result['p_model'] * 100:.1f}%")
    print(f"Book Prob:   {result['p_book'] * 100:.1f}%")
    print(f"Model Projection: {result['projection']:.1f} {stat}")
    print(f"EV: {result['ev'] * 100:.1f}¢ per $1 | {'🔥 Positive' if result['ev'] > 0 else '⚠️ Negative'}")
    print("🟢 Over Value" if result['projection'] > line else "🔴 Under Value")
    print("="*60 + "\n")
    
    return result


if __name__ == "__main__":
    print("🧠 PropPulse+ Model v2025.3 — Player Prop EV Analyzer")
    print("=" * 60)

    if len(sys.argv) > 1 and sys.argv[1].endswith(".csv"):
        csv_path = sys.argv[1]
        if os.path.exists(csv_path):
            run_csv_batch_mode(csv_path)
        else:
            print(f"❌ CSV not found: {csv_path}")
            sys.exit(1)
    else:
        print("\n💡 Entering interactive mode...\n")
        try:
            main()
        except KeyboardInterrupt:
            print("\n⚠️ Interrupted by user. Exiting...")
            sys.exit(0)
        except Exception as e:
            print(f"\n❌ Fatal error: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)



